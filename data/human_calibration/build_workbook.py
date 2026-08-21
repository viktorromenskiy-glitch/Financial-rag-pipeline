"""Builds the human-calibration labeling workbook (Plan-2, item 4).

Reads labeling_sample.json (30 questions, stratified by source_dataset,
proportional to the n=250 corpus mix, fixed seed for reproducibility) and
writes human_calibration_labeling.xlsx for Viktor to fill in by hand.

Deliberately blind: the workbook does NOT include the judge's verdict or
the deterministic_match result anywhere - those live only in
answer_key.json, which is not sent to the user. Showing the judge's
verdict during labeling would anchor the human judgment on it, defeating
the point of an independent calibration check.
"""

import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

with open("labeling_sample.json", "r", encoding="utf-8") as f:
    rows = json.load(f)

wb = Workbook()

# --- Sheet 1: Instructions ------------------------------------------------
ws_instr = wb.active
ws_instr.title = "Инструкция"
ws_instr.sheet_view.showGridLines = False
ws_instr.column_dimensions["A"].width = 110

# Explicit black (not "automatic"/None) - some viewers (Excel/Sheets in
# dark theme, some mobile apps) render an unset "automatic" font color as
# white, which is invisible against the yellow input-cell fill. This was
# reported by Viktor (entries didn't appear to show up) and is the most
# likely cause - an explicit color removes the ambiguity everywhere.
BLACK = "FF000000"
title_font = Font(name="Arial", size=14, bold=True, color=BLACK)
body_font = Font(name="Arial", size=11, color=BLACK)
bold_font = Font(name="Arial", size=11, bold=True, color=BLACK)

lines = [
    ("Калибровка LLM-судьи на человеческой разметке (план доработки-2, пункт 4)", title_font),
    ("", body_font),
    (
        "Зачем: итоговые цифры проекта (76.0-76.8% точности) целиком опираются на вердикт "
        "LLM-судьи (Claude Sonnet 5) плюс детерминированную числовую проверку. Судья ни разу "
        "не сверялся с реальной человеческой разметкой - только с автоматическим числовым "
        "чекером (93.3% согласия). Это калибровочное исследование закрывает этот пробел, по "
        "методологии, использованной в одном из проектов-конкурентов "
        "(joaopaulotr/financebench-rag-eval) - и находкам, которые тот проект показал: судья "
        "может систематически завышать точность из-за 'fluency bias' (беглый, убедительный "
        "ответ с неверным числом ошибочно засчитывается как верный).",
        body_font,
    ),
    ("", body_font),
    (
        "Что делать: на листе 'Разметка' — 30 вопросов, выбранных случайно (фиксированный "
        "seed) из прогона results/error_analysis_250, пропорционально распределению по "
        "источникам (TAT-DQA / FinQA / ConvFinQA). Для каждой строки в столбце F "
        "(с рамкой) укажите CORRECT, если 'Ответ модели' по существу совпадает с "
        "'Эталонным ответом', или INCORRECT — если нет.",
        body_font,
    ),
    ("", body_font),
    ("Критерий совпадения — тот же, что используется в промпте LLM-судьи (дословно):", bold_font),
    (
        "Засчитывайте как CORRECT: небольшие расхождения в округлении; разница в знаке, если "
        "направление неоднозначно из вопроса (например, -60 против 60); эквивалентное "
        "представление в процентах вместо доли (1.5 против 0.015); эквивалентное представление "
        "в разных единицах одного и того же значения (например, 5413606 против 5413606000, "
        "если одно — в тысячах, а другое — в абсолютных единицах).",
        body_font,
    ),
    (
        "Это не означает 'засчитывать всё подряд' — если число по существу другое (не форма "
        "записи, а другое значение), это INCORRECT, даже если формулировка ответа звучит "
        "убедительно.",
        body_font,
    ),
    ("", body_font),
    (
        "Важно: разметка должна быть слепой. В файле нет вердикта судьи и нет результата "
        "детерминированной проверки — специально, чтобы ваша оценка была независимой, а не "
        "подстроенной под то, что 'сказала' модель. Пожалуйста, не открывайте "
        "results/error_analysis_250/eval_results.jsonl и не ищите question_id по репозиторию, "
        "пока не закончите разметку всех 30 строк.",
        body_font,
    ),
    ("", body_font),
    (
        "Когда закончите: сохраните файл и пришлите его обратно (или просто скажите, что "
        "разметка готова, если работаете в этом же диалоге) - я посчитаю согласие/несогласие "
        "с судьёй, TPR/TNR и задокументирую результат в ТЗ и README, как договаривались.",
        body_font,
    ),
    ("", body_font),
    ("Пример (не входит в 30 — просто показывает формат столбца F):", bold_font),
]

r = 1
for text, font in lines:
    cell = ws_instr.cell(row=r, column=1, value=text)
    cell.font = font
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if text.startswith("Калибровка"):
        ws_instr.row_dimensions[r].height = 22
    elif text:
        ws_instr.row_dimensions[r].height = 60
    r += 1

# Small example table on the instructions sheet
ex_header_row = r + 1
headers = ["#", "Датасет", "Вопрос", "Ответ модели", "Эталонный ответ", "Ваш вердикт"]
grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
for c, h in enumerate(headers, start=1):
    cell = ws_instr.cell(row=ex_header_row, column=c, value=h)
    cell.font = bold_font
    cell.fill = grey_fill
example_row = ex_header_row + 1
example_values = [
    "EX",
    "FinQA",
    "What was the change in revenue from 2017 to 2018, in millions?",
    "42",
    "42.3",
    "CORRECT",
]
example_border = Border(left=Side(style="medium", color=BLACK), right=Side(style="medium", color=BLACK),
                         top=Side(style="medium", color=BLACK), bottom=Side(style="medium", color=BLACK))
for c, v in enumerate(example_values, start=1):
    cell = ws_instr.cell(row=example_row, column=c, value=v)
    cell.font = body_font
ws_instr.cell(row=example_row, column=6).border = example_border
ws_instr.cell(
    row=example_row + 1,
    column=1,
    value=(
        "^ пример: 42 против 42.3 - небольшое расхождение в округлении, засчитано как CORRECT "
        "по тому же правилу, что использует судья."
    ),
).font = Font(name="Arial", size=10, italic=True, color=BLACK)
ws_instr.row_dimensions[example_row + 1].height = 30

# --- Sheet 2: Labeling -----------------------------------------------------
ws = wb.create_sheet("Разметка")
headers = ["#", "Датасет", "Вопрос", "Ответ модели", "Эталонный ответ", "Ваш вердикт (CORRECT/INCORRECT)", "Комментарий (опционально)"]
widths = [5, 12, 60, 22, 22, 30, 30]
for c, (h, w) in enumerate(zip(headers, widths), start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = bold_font
    cell.fill = grey_fill
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions[chr(64 + c)].width = w
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

# No fill on the input column any more - Viktor reported that even with
# explicit black+bold text, the yellow fill made his entries unreadable in
# whatever app he's using to edit (color-on-color rendering issue outside
# our control). A border marks the input cell instead of a fill, which
# can't visually compete with the text drawn on top of it.
input_border = Border(left=Side(style="medium", color=BLACK), right=Side(style="medium", color=BLACK),
                       top=Side(style="medium", color=BLACK), bottom=Side(style="medium", color=BLACK))

for i, row in enumerate(rows, start=2):
    ws.cell(row=i, column=1, value=row["label_id"]).font = body_font
    ws.cell(row=i, column=2, value=row["source_dataset"]).font = body_font
    qc = ws.cell(row=i, column=3, value=row["question"])
    qc.font = body_font
    qc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=i, column=4, value=row["generated_answer"]).font = body_font
    ws.cell(row=i, column=5, value=row["gold_answer"]).font = body_font
    verdict_cell = ws.cell(row=i, column=6)
    verdict_cell.border = input_border
    verdict_cell.font = Font(name="Arial", size=12, bold=True, color=BLACK)
    ws.row_dimensions[i].height = 45

dv = DataValidation(type="list", formula1='"CORRECT,INCORRECT"', allow_blank=True, showDropDown=False)
dv.error = "Выберите CORRECT или INCORRECT из списка"
dv.errorTitle = "Неверное значение"
ws.add_data_validation(dv)
dv.add(f"F2:F{len(rows) + 1}")

wb.save("human_calibration_labeling.xlsx")
print("wrote human_calibration_labeling.xlsx,", len(rows), "rows")
