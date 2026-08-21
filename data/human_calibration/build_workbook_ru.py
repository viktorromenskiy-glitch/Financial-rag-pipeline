"""Builds a Russian-language READ-ALONG reference for the human-calibration
labeling task (plan doradotki-2, item 4).

Viktor asked to fill in the verdict column in the original
human_calibration_labeling.xlsx (that file stays the source of record),
but reads English only with difficulty and wants the question text in
Russian on a second monitor while he labels.

This file is a reference only - it has NO verdict column, so there is no
risk of the two files disagreeing about what was actually entered. Row
numbers (#) match human_calibration_labeling.xlsx exactly, 1-30, so a
row here lines up with the same row there. Only the question text is
translated; generated_answer/gold_answer are left as-is (they are numeric
strings, not natural language - translating them would only add risk of
a transcription error, not add information). Still no judge verdict and
no deterministic_match anywhere in this file, same as the English one -
translating for readability doesn't change the blind-labeling requirement.
"""

import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

with open("labeling_sample.json", "r", encoding="utf-8") as f:
    rows = json.load(f)
with open("translations_ru.json", "r", encoding="utf-8") as f:
    translations = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = "Разметка (RU, справка)"
ws.sheet_view.showGridLines = True

# Explicit black - see build_workbook.py for why (an unset "automatic"
# color can render as white in a dark-themed viewer, invisible against a
# colored fill).
BLACK = "FF000000"
body_font = Font(name="Arial", size=11, color=BLACK)
bold_font = Font(name="Arial", size=11, bold=True, color=BLACK)
italic_font = Font(name="Arial", size=10, italic=True, color=BLACK)
grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

note = (
    "Это справочный файл только для чтения - вписывать вердикт нужно в "
    "human_calibration_labeling.xlsx (столбец F), не сюда. Номера строк (#) "
    "здесь совпадают со строками в том файле 1-в-1. Вопрос переведён на "
    "русский язык (машинный перевод для удобства чтения); ответ модели и "
    "эталонный ответ оставлены как есть - это числа, не текст, переводить "
    "нечего."
)
cell = ws.cell(row=1, column=1, value=note)
cell.font = italic_font
cell.alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
ws.row_dimensions[1].height = 45

headers = ["#", "Датасет", "Вопрос (перевод на русский)", "Ответ модели", "Эталонный ответ"]
widths = [5, 12, 75, 22, 22]
header_row = 2
for c, (h, w) in enumerate(zip(headers, widths), start=1):
    cell = ws.cell(row=header_row, column=c, value=h)
    cell.font = bold_font
    cell.fill = grey_fill
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions[chr(64 + c)].width = w
ws.row_dimensions[header_row].height = 24
ws.freeze_panes = f"A{header_row + 1}"

r = header_row + 1
for row in rows:
    label_id = row["label_id"]
    ws.cell(row=r, column=1, value=label_id).font = body_font
    ws.cell(row=r, column=2, value=row["source_dataset"]).font = body_font
    qc = ws.cell(row=r, column=3, value=translations[str(label_id)])
    qc.font = body_font
    qc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=4, value=row["generated_answer"]).font = body_font
    ws.cell(row=r, column=5, value=row["gold_answer"]).font = body_font
    ws.row_dimensions[r].height = 45
    if row["generated_answer"] == "INSUFFICIENT_CONTEXT":
        note_cell = ws.cell(
            row=r,
            column=6,
            value=(
                "INSUFFICIENT_CONTEXT — служебное значение: модель посчитала, что "
                "найденного контекста недостаточно для ответа, и вернула эту метку "
                "вместо числа (не гадала)."
            ),
        )
        note_cell.font = italic_font
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions["F"].width = 45
    r += 1

wb.save("human_calibration_labeling_RU_reference.xlsx")
print("wrote human_calibration_labeling_RU_reference.xlsx,", len(rows), "rows")
